from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


def write_rename_voc(path_to_file, image_info, shoot_id):
    # Проверяем существование файла, создаем новый, если он отсутствует
    if not Path(path_to_file).is_file():
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet_name"
        wb.save(path_to_file)

    wb = load_workbook(path_to_file)

    # Создаем новый лист с заданным именем `shoot_id`
    if shoot_id in wb.sheetnames:
        ws = wb[shoot_id]
    else:
        ws = wb.create_sheet(shoot_id)

    # Устанавливаем ширину колонок
    column_widths = {'A': 18, 'B': 18, 'C': 8, 'D': 70}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Заголовки таблицы
    headers = ["KSP name", "original file name", "published", "caption"]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    # Добавляем данные
    last_line = ws.max_row
    red_font = Font(color="FF0000")
    centered_text = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for k, values in image_info.items():
        last_line += 1
        ws.cell(row=last_line, column=1, value=k).font = red_font if len(values) >= 2 else None
        ws.cell(row=last_line, column=2, value=values[0]).font = red_font if len(values) >= 2 else None
        if len(values) >= 2:
            ws.cell(row=last_line, column=3, value=values[1]).font = red_font
            caption_cell = ws.cell(row=last_line, column=4, value=values[2])
            caption_cell.font = red_font
            caption_cell.alignment = centered_text

    # Сохраняем изменения
    wb.save(path_to_file)
